import pandas as pd
from pathlib import Path
from openpyxl import Workbook

class Armazenamento:
    
    """Gerencia o salvamento e leitura de dados em um único arquivo Excel."""

    def __init__(self, nome_arquivo="dados.xlsx"):

        self.caminho = Path(nome_arquivo)

        # Cria o arquivo com as abas se ainda não existir
        if not self.caminho.exists():
            self._criar_arquivo_inicial()

    def _criar_arquivo_inicial(self):

        """Cria um novo arquivo Excel com abas para cada tipo."""

        wb = Workbook()
        abas = ["pacientes", "pesquisadores", "fisioterapeutas"]

        # Remove a aba padrão criada automaticamente
        padrao = wb.active
        wb.remove(padrao)

        for aba in abas:
            ws = wb.create_sheet(title=aba)
            ws.append(["ID", "Nome"])

        wb.save(self.caminho)

    def _verificar_aba(self, aba):

        """Verifica se a aba existe, senão cria."""

        from openpyxl import load_workbook
        wb = load_workbook(self.caminho)

        if aba not in wb.sheetnames:
            ws = wb.create_sheet(title=aba)
            ws.append(["ID", "Nome"])
            wb.save(self.caminho)
        wb.close()

    def salvar(self, aba, dados):

        """Salva um novo registro na aba correspondente."""

        self._verificar_aba(aba)
        try:
            df_existente = pd.read_excel(self.caminho, sheet_name=aba)
        except ValueError:
            df_existente = pd.DataFrame(columns=dados.keys())

        novo_df = pd.DataFrame([dados])
        df = pd.concat([df_existente, novo_df], ignore_index=True)

        with pd.ExcelWriter(self.caminho, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=aba, index=False)

    def carregar(self, aba):

        """Carrega os dados de uma aba específica."""
        
        self._verificar_aba(aba)
        return pd.read_excel(self.caminho, sheet_name=aba)
